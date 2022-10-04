import inspect
import logging
import configparser
import softest
import openpyxl
from openpyxl import Workbook, load_workbook

config = configparser.RawConfigParser()
config.read("C:/Users/Jyothibabu/PycharmProjects/nopcommerceapp/configurations/config.ini")


class Utils(softest.TestCase):

    # ************** Start Read Config Data **************
    @staticmethod
    def getApplicationURL():
        url = config.get('common info', 'baseURL')
        return url

    @staticmethod
    def getUsername():
        username = config.get('common info', 'username')
        return username

    @staticmethod
    def getPassword():
        password = config.get('common info', 'password')
        return password

    # *************    Start Logging     ***********
    def custom_logger(logLevel=logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("logs/automation.log", mode='a')
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    # ***************  Read Excel Test Data  **************
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return(sheet.max_row)

    def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return(sheet.max_column)

    def readData(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)